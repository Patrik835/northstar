import csv
import hashlib
import io
import re
import unicodedata
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.encryption import CredentialCipher
from app.integrations.connectors.base import ConnectorPosition
from app.integrations.market_data import EcbFxRateProvider, FxRateError, FxRateProvider
from app.models.broker import BrokerConnection
from app.models.enums import AssetType, Broker, ConnectionStatus, TransactionType
from app.models.portfolio import PortfolioSnapshot, Position, Transaction
from app.repositories.connections import ConnectionRepository
from app.services.instrument_resolver import InstrumentResolver

MAX_XTB_FILE_BYTES = 15 * 1024 * 1024
MAX_XTB_UPLOAD_BYTES = 30 * 1024 * 1024
MAX_XTB_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024


class XtbImportError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class XtbUpload:
    filename: str
    content: bytes


@dataclass(frozen=True, slots=True)
class ParsedXtbPosition:
    instrument_id: str
    ticker: str
    canonical_symbol: str
    name: str | None
    isin: str | None
    asset_type: AssetType
    quantity: Decimal
    average_price: Decimal | None
    current_value: Decimal
    currency: str
    reported_pnl: Decimal | None


@dataclass(frozen=True, slots=True)
class ParsedXtbTransaction:
    external_id: str
    ticker: str
    transaction_type: TransactionType
    quantity: Decimal | None
    price: Decimal | None
    value: Decimal
    currency: str
    executed_at: datetime


@dataclass(frozen=True, slots=True)
class ParsedXtbFile:
    rows_read: int
    positions: list[ParsedXtbPosition]
    transactions: list[ParsedXtbTransaction]
    has_position_snapshot: bool
    valued_at: datetime
    warnings: list[str]
    limitation_warnings: list[str]


@dataclass(frozen=True, slots=True)
class XtbImportSummary:
    connection: BrokerConnection
    rows_read: int
    transactions_added: int
    duplicates_skipped: int
    positions_imported: int
    warnings: list[str]


HEADER_ALIASES = {
    "id": {"id", "order", "orderid", "position", "positionid", "dealid"},
    "symbol": ("ticker", "symbol", "instrument", "instrumentname", "market"),
    "instrument_position": {"instrumentposition"},
    "name": {"name", "description", "instrumentdescription", "company"},
    "isin": {"isin", "isincode"},
    "asset_type": {
        "assetclass",
        "assettype",
        "category",
        "instrumentclass",
        "instrumenttype",
        "producttype",
    },
    "operation": {"action", "direction", "operation", "transactiontype", "type"},
    "status": {"state", "status"},
    "quantity": {
        "amount",
        "numberofshares",
        "numberofunits",
        "quantity",
        "shares",
        "units",
        "volume",
    },
    "open_price": {
        "averageprice",
        "bookprice",
        "entryprice",
        "openprice",
        "openingprice",
        "purchaseprice",
    },
    "current_price": {"closeprice", "currentprice", "marketprice", "price"},
    "current_value": {
        "amount",
        "currentvalue",
        "marketvalue",
        "positionvalue",
        "value",
    },
    "purchase_value": {"purchasevalue"},
    "sale_value": {"salevalue"},
    "profit": {
        "gainloss",
        "netprofit",
        "performance",
        "pl",
        "profit",
        "profitloss",
        "result",
    },
    "currency": {"accountcurrency", "currency", "curr", "mena", "waluta"},
    "date": {"date", "datetime", "time", "timestamp", "transactiondate"},
    "open_time": {
        "entrytime",
        "opendate",
        "opentime",
        "opentimeutc",
        "openingtime",
    },
    "close_time": {"closedate", "closetime", "closetimeutc", "closingtime"},
    "commission": {"commission", "fee", "fees"},
    "swap": {"swap", "swapvalue"},
}

_UNSUPPORTED_ASSETS = ("cfd", "commodity", "forex", "fx", "index", "crypto")
_KNOWN_CURRENCIES = {
    "AUD",
    "BGN",
    "BRL",
    "CAD",
    "CHF",
    "CNY",
    "CZK",
    "DKK",
    "EUR",
    "GBP",
    "HKD",
    "HUF",
    "IDR",
    "ILS",
    "INR",
    "ISK",
    "JPY",
    "KRW",
    "MXN",
    "MYR",
    "NOK",
    "NZD",
    "PHP",
    "PLN",
    "RON",
    "SEK",
    "SGD",
    "THB",
    "TRY",
    "USD",
    "ZAR",
}
_EXCHANGE_SUFFIX = re.compile(
    r"\.(AT|BE|CH|CZ|DE|DK|ES|FI|FR|IT|NL|NO|PL|PT|SE|SK|UK|US)$",
    re.IGNORECASE,
)


class XtbStatementParser:
    """Parse XTB's regional CSV and Excel exports without relying on one fixed layout."""

    def parse(self, upload: XtbUpload) -> ParsedXtbFile:
        if not upload.content:
            raise XtbImportError(f"{upload.filename} is empty.")
        if len(upload.content) > MAX_XTB_FILE_BYTES:
            raise XtbImportError(f"{upload.filename} is larger than the 15 MB limit.")

        suffix = Path(upload.filename).suffix.lower()
        if suffix == ".csv":
            sheets = [(Path(upload.filename).stem, self._csv_rows(upload.content))]
        elif suffix == ".xlsx":
            sheets = self._excel_rows(upload.content)
        else:
            raise XtbImportError("XTB imports support CSV and XLSX files.")

        rows_read = 0
        positions: list[ParsedXtbPosition] = []
        transactions: list[ParsedXtbTransaction] = []
        warnings: list[str] = []
        limitations: list[str] = []
        has_snapshot = False
        valued_at: datetime | None = None
        found_table = False

        for sheet_name, rows in sheets:
            sheet_valued_at = _latest_date(rows) or datetime.now(timezone.utc)
            default_currency = _report_currency(rows, upload.filename) or "EUR"
            cash = _cash_position(rows)
            if cash is not None and not any(
                item.asset_type is AssetType.CASH for item in positions
            ):
                positions.append(cash)
            for section, headers, data_rows in self._tables(rows, sheet_name):
                found_table = True
                fields = _resolve_fields(headers)
                position_table = "symbol" in fields and "quantity" in fields
                explicit_live_value = _normalize_header(fields.get("current_value", "")) in {
                    "currentvalue",
                    "marketvalue",
                    "positionvalue",
                }
                explicit_live_price = _normalize_header(fields.get("current_price", "")) in {
                    "currentprice",
                    "marketprice",
                }
                snapshot_table = (
                    position_table
                    and "close_time" not in fields
                    and (
                        "openposition" in _normalize_header(section)
                        or "open_time" in fields
                        or explicit_live_value
                        or explicit_live_price
                    )
                )
                has_snapshot = has_snapshot or snapshot_table
                if snapshot_table and (valued_at is None or sheet_valued_at > valued_at):
                    valued_at = sheet_valued_at

                for row_number, values in data_rows:
                    if not any(str(value or "").strip() for value in values):
                        continue
                    rows_read += 1
                    row = {
                        header: values[index] if index < len(values) else None
                        for index, header in enumerate(headers)
                    }
                    try:
                        parsed_positions, parsed_transactions, row_warnings = self._parse_row(
                            row, fields, section, snapshot_table, default_currency
                        )
                    except XtbImportError as exc:
                        raise XtbImportError(f"{upload.filename}, row {row_number}: {exc}") from exc
                    positions.extend(parsed_positions)
                    transactions.extend(parsed_transactions)
                    warnings.extend(row_warnings)
                    limitations.extend(
                        warning for warning in row_warnings if warning.startswith("Skipped ")
                    )

        if not found_table:
            raise XtbImportError(
                f"{upload.filename} does not contain a recognized XTB positions or history table."
            )
        if not positions and not transactions and not has_snapshot:
            raise XtbImportError(f"{upload.filename} contains no supported XTB investment data.")

        return ParsedXtbFile(
            rows_read=rows_read,
            positions=positions,
            transactions=_deduplicate_transactions(transactions),
            has_position_snapshot=has_snapshot,
            valued_at=valued_at or datetime.now(timezone.utc),
            warnings=_unique(warnings),
            limitation_warnings=_unique(limitations),
        )

    @staticmethod
    def _csv_rows(content: bytes) -> list[list[object]]:
        text = None
        for encoding in ("utf-8-sig", "utf-16", "cp1250", "cp1252"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise XtbImportError("The CSV text encoding is not supported.")
        try:
            dialect = csv.Sniffer().sniff(text[:16384], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return [list(row) for row in csv.reader(io.StringIO(text), dialect=dialect)]

    @staticmethod
    def _excel_rows(content: bytes) -> list[tuple[str, list[list[object]]]]:
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                if (
                    sum(item.file_size for item in archive.infolist())
                    > MAX_XTB_XLSX_UNCOMPRESSED_BYTES
                ):
                    raise XtbImportError("The XLSX expands beyond the 100 MB safety limit.")
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except XtbImportError:
            raise
        except Exception as exc:
            raise XtbImportError("The XLSX file could not be read.") from exc
        try:
            result = []
            for sheet in workbook.worksheets:
                # Native XTB workbooks currently declare A1:A1 even when the worksheet
                # contains a complete report. ReadOnlyWorksheet otherwise stops after the
                # account-number cell and never reaches the table headers.
                sheet.reset_dimensions()
                result.append(
                    (sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
                )
            return result
        finally:
            workbook.close()

    @staticmethod
    def _tables(
        rows: list[list[object]], sheet_name: str
    ) -> list[tuple[str, list[str], list[tuple[int, list[object]]]]]:
        tables: list[tuple[str, list[str], list[tuple[int, list[object]]]]] = []
        starts: list[tuple[int, list[str]]] = []
        for index, row in enumerate(rows):
            headers = [str(value or "").strip() for value in row]
            fields = _resolve_fields(headers)
            position_header = "symbol" in fields and "quantity" in fields
            activity_header = (
                ("date" in fields or "open_time" in fields or "close_time" in fields)
                and "operation" in fields
                and ("current_value" in fields or "profit" in fields or "symbol" in fields)
            )
            if position_header or activity_header:
                starts.append((index, headers))

        for start_index, headers in starts:
            next_starts = [index for index, _ in starts if index > start_index]
            end_index = min(next_starts) if next_starts else len(rows)
            section_parts = [sheet_name]
            for previous in rows[max(0, start_index - 3) : start_index]:
                values = [
                    str(value or "").strip() for value in previous if str(value or "").strip()
                ]
                if len(values) == 1:
                    section_parts.append(values[0])
            data = [(index + 1, rows[index]) for index in range(start_index + 1, end_index)]
            tables.append((" ".join(section_parts), headers, data))
        return tables

    def _parse_row(
        self,
        row: dict[str, object],
        fields: dict[str, str],
        section: str,
        snapshot_table: bool,
        default_currency: str,
    ) -> tuple[list[ParsedXtbPosition], list[ParsedXtbTransaction], list[str]]:
        symbol = _value(row, fields, "symbol").strip().upper()
        operation = _value(row, fields, "operation").strip().lower()
        status = _value(row, fields, "status").strip().lower()
        asset_text = " ".join(filter(None, (_value(row, fields, "asset_type"), section))).lower()
        if _transaction_type(operation) is None:
            asset_text = f"{asset_text} {operation}"

        asset_type = _asset_type(asset_text)
        unsupported = "cfd" in asset_text or (
            asset_type is AssetType.OTHER
            and any(token in asset_text for token in _UNSUPPORTED_ASSETS)
        )
        if unsupported:
            label = symbol or "an unsupported product"
            return [], [], [f"Skipped {label} because this connector imports stocks and ETFs only."]
        if status and any(token in status for token in ("cancel", "pending", "reject")):
            return [], [], []

        quantity = _decimal(_value(row, fields, "quantity"))
        open_price = _decimal(_value(row, fields, "open_price"))
        current_price = _decimal(_value(row, fields, "current_price"))
        current_value = _decimal(_value(row, fields, "current_value"))
        profit = _decimal(_value(row, fields, "profit"))
        currency = _row_currency(row, fields) or default_currency
        isin = _clean_isin(_value(row, fields, "isin"))
        instrument_position = _value(row, fields, "instrument_position").strip()
        name = _value(row, fields, "name").strip() or None
        row_id = _value(row, fields, "id").strip()
        open_time = _optional_datetime(_value(row, fields, "open_time"))
        close_time = _optional_datetime(_value(row, fields, "close_time"))
        generic_time = _optional_datetime(_value(row, fields, "date"))
        canonical_symbol = _canonical_symbol(symbol)
        warnings: list[str] = []
        positions: list[ParsedXtbPosition] = []
        transactions: list[ParsedXtbTransaction] = []

        hierarchical_positions = "instrument_position" in fields
        summary_position = not hierarchical_positions or (
            asset_type in {AssetType.STOCK, AssetType.ETF} and not operation
        )
        if hierarchical_positions:
            if summary_position:
                name = name or instrument_position or None
            else:
                row_id = row_id or instrument_position

        is_open = (
            snapshot_table
            and summary_position
            and not close_time
            and not any(token in status for token in ("closed", "close"))
        )
        if is_open and symbol and quantity is not None and quantity != 0:
            if current_value is None and current_price is not None:
                current_value = abs(quantity) * abs(current_price)
            if current_value is None:
                raise XtbImportError(f"{symbol} has no current price or market value")
            if asset_type is AssetType.OTHER:
                warnings.append(
                    f"XTB did not label {symbol} as a stock or ETF; it was imported as Other."
                )
            if profit is None and open_price is not None and current_price is not None:
                profit = abs(quantity) * (current_price - open_price)
            provider_id = isin or row_id or f"XTB:{symbol}"
            positions.append(
                ParsedXtbPosition(
                    instrument_id=provider_id[:120],
                    ticker=symbol,
                    canonical_symbol=canonical_symbol,
                    name=name or symbol,
                    isin=isin,
                    asset_type=asset_type,
                    quantity=abs(quantity),
                    average_price=abs(open_price) if open_price is not None else None,
                    current_value=abs(current_value),
                    currency=currency,
                    reported_pnl=profit,
                )
            )

        if symbol and quantity is not None and open_price is not None and open_time:
            purchase_value = _decimal(_value(row, fields, "purchase_value"))
            transactions.append(
                _transaction(
                    row_id=row_id,
                    action="open",
                    ticker=symbol,
                    transaction_type=TransactionType.BUY,
                    quantity=abs(quantity),
                    price=abs(open_price),
                    value=abs(purchase_value or quantity * open_price),
                    currency=currency,
                    executed_at=open_time,
                )
            )
        if symbol and quantity is not None and current_price is not None and close_time:
            sale_value = _decimal(_value(row, fields, "sale_value"))
            transactions.append(
                _transaction(
                    row_id=row_id,
                    action="close",
                    ticker=symbol,
                    transaction_type=TransactionType.SELL,
                    quantity=abs(quantity),
                    price=abs(current_price),
                    value=abs(sale_value or quantity * current_price),
                    currency=currency,
                    executed_at=close_time,
                )
            )

        direct_type = _transaction_type(operation)
        transaction_time = generic_time or open_time or close_time
        if direct_type and transaction_time and not transactions:
            value = current_value if current_value is not None else profit
            if value is None and quantity is not None and current_price is not None:
                value = quantity * current_price
            if value is not None:
                transactions.append(
                    _transaction(
                        row_id=row_id,
                        action=direct_type.value,
                        ticker=symbol or currency,
                        transaction_type=direct_type,
                        quantity=(
                            abs(quantity)
                            if quantity is not None
                            and direct_type in {TransactionType.BUY, TransactionType.SELL}
                            else None
                        ),
                        price=abs(current_price) if current_price is not None else None,
                        value=abs(value),
                        currency=currency,
                        executed_at=transaction_time,
                    )
                )

        fee_time = close_time or open_time or generic_time
        for fee_name in ("commission", "swap"):
            fee = _decimal(_value(row, fields, fee_name))
            if fee and fee_time:
                transactions.append(
                    _transaction(
                        row_id=row_id,
                        action=fee_name,
                        ticker=symbol or currency,
                        transaction_type=TransactionType.FEE,
                        quantity=None,
                        price=None,
                        value=abs(fee),
                        currency=currency,
                        executed_at=fee_time,
                    )
                )
        return positions, transactions, warnings


class XtbImportService:
    def __init__(
        self,
        db: AsyncSession,
        cipher: CredentialCipher,
        fx_rates: FxRateProvider | None = None,
    ) -> None:
        self.db = db
        self.cipher = cipher
        self.fx_rates = fx_rates or EcbFxRateProvider(db)

    async def import_files(self, user_id: uuid.UUID, uploads: list[XtbUpload]) -> XtbImportSummary:
        if not uploads:
            raise XtbImportError("Select at least one XTB CSV or XLSX report.")
        if sum(len(upload.content) for upload in uploads) > MAX_XTB_UPLOAD_BYTES:
            raise XtbImportError("The combined XTB upload is larger than 30 MB.")

        parsed_files = [XtbStatementParser().parse(upload) for upload in uploads]
        snapshot_files = [item for item in parsed_files if item.has_position_snapshot]
        snapshot = max(snapshot_files, key=lambda item: item.valued_at) if snapshot_files else None
        transactions = _deduplicate_transactions(
            [transaction for item in parsed_files for transaction in item.transactions]
        )
        warnings = _unique([warning for item in parsed_files for warning in item.warnings])
        limitations = _unique(
            [warning for item in parsed_files for warning in item.limitation_warnings]
        )

        connection = await ConnectionRepository(self.db).by_broker(user_id, Broker.XTB)
        if connection is None:
            connection = BrokerConnection(
                user_id=user_id,
                broker=Broker.XTB,
                encrypted_credentials=self.cipher.encrypt({"source": "statement"}),
                credential_hint="File import",
                status=ConnectionStatus.PENDING,
            )
            self.db.add(connection)
            await self.db.flush()

        existing_ids = set(
            await self.db.scalars(
                select(Transaction.external_id).where(
                    Transaction.broker_connection_id == connection.id
                )
            )
        )
        new_transactions = [item for item in transactions if item.external_id not in existing_ids]
        transaction_values_eur: list[Decimal | None] = []
        historical_fx_missing = False
        for item in new_transactions:
            try:
                transaction_values_eur.append(
                    await self.fx_rates.convert_to_eur(
                        item.value, item.currency, item.executed_at.date()
                    )
                )
            except FxRateError:
                # Historical rates are accumulated from deployment onward. Preserve the
                # exact native amount rather than assigning an incorrect current FX rate.
                transaction_values_eur.append(None)
                historical_fx_missing = True
        if historical_fx_missing:
            warnings.append(
                "Some historical activity has no matching stored ECB rate yet; "
                "its original-currency value was preserved without an estimated EUR value."
            )

        self.db.add_all(
            [
                Transaction(
                    broker_connection_id=connection.id,
                    external_id=item.external_id,
                    ticker=item.ticker,
                    transaction_type=item.transaction_type,
                    quantity=item.quantity,
                    price=item.price,
                    value=item.value,
                    value_eur=value_eur,
                    currency=item.currency,
                    executed_at=item.executed_at,
                )
                for item, value_eur in zip(new_transactions, transaction_values_eur, strict=True)
            ]
        )

        positions_imported = 0
        if snapshot is not None:
            grouped = _aggregate_positions(snapshot.positions)
            try:
                values_eur = [
                    await self.fx_rates.convert_to_eur(item.current_value, item.currency)
                    for item in grouped
                ]
                pnl_eur = [
                    None
                    if item.reported_pnl is None
                    else await self.fx_rates.convert_to_eur(item.reported_pnl, item.currency)
                    for item in grouped
                ]
            except FxRateError as exc:
                await self.db.rollback()
                raise XtbImportError(str(exc)) from exc

            resolver = InstrumentResolver(self.db)
            canonical = []
            for item in grouped:
                canonical.append(
                    await resolver.resolve(
                        Broker.XTB,
                        ConnectorPosition(
                            instrument_id=item.instrument_id,
                            ticker=item.ticker,
                            name=item.name,
                            asset_type=item.asset_type,
                            quantity=item.quantity,
                            average_price=item.average_price,
                            current_value=item.current_value,
                            currency=item.currency,
                            canonical_symbol=item.canonical_symbol,
                            isin=item.isin,
                            reported_pnl=item.reported_pnl,
                        ),
                    )
                )

            await self.db.execute(
                delete(Position).where(Position.broker_connection_id == connection.id)
            )
            self.db.add_all(
                [
                    Position(
                        broker_connection_id=connection.id,
                        instrument_id=item.instrument_id,
                        canonical_instrument_id=instrument.id,
                        ticker=item.ticker,
                        name=item.name,
                        asset_type=instrument.asset_type,
                        quantity=item.quantity,
                        average_price=item.average_price,
                        current_value=item.current_value,
                        currency=item.currency,
                        current_value_eur=value_eur,
                        reported_pnl=item.reported_pnl,
                        reported_pnl_eur=item_pnl_eur,
                        valued_at=snapshot.valued_at,
                        valuation_source="statement",
                        is_estimated=False,
                    )
                    for item, value_eur, item_pnl_eur, instrument in zip(
                        grouped, values_eur, pnl_eur, canonical, strict=True
                    )
                ]
            )
            total_value_eur = sum(values_eur, Decimal(0))
            reported_pnl_values = [value for value in pnl_eur if value is not None]
            total_pnl_eur = sum(reported_pnl_values, Decimal(0)) if reported_pnl_values else None
            stored_snapshot = await self.db.scalar(
                select(PortfolioSnapshot).where(
                    PortfolioSnapshot.broker_connection_id == connection.id,
                    PortfolioSnapshot.snapshot_date == snapshot.valued_at.date(),
                )
            )
            if stored_snapshot:
                stored_snapshot.total_value = total_value_eur
                stored_snapshot.total_value_eur = total_value_eur
                stored_snapshot.currency = "EUR"
                stored_snapshot.reported_pnl = total_pnl_eur
                stored_snapshot.reported_pnl_eur = total_pnl_eur
            else:
                self.db.add(
                    PortfolioSnapshot(
                        broker_connection_id=connection.id,
                        snapshot_date=snapshot.valued_at.date(),
                        total_value=total_value_eur,
                        total_value_eur=total_value_eur,
                        currency="EUR",
                        reported_pnl=total_pnl_eur,
                        reported_pnl_eur=total_pnl_eur,
                    )
                )
            positions_imported = len(grouped)

        completed_at = datetime.now(timezone.utc)
        connection.status = ConnectionStatus.LIMITED if limitations else ConnectionStatus.ACTIVE
        connection.last_error = " ".join(limitations[:3]) or None
        connection.last_sync_attempt_at = completed_at
        connection.last_successful_sync_at = completed_at
        connection.last_synced_at = completed_at
        connection.reconciliation_difference_percent = None
        connection.reconciliation_checked_at = None
        connection.reconciliation_warning = None
        await self.db.commit()
        await self.db.refresh(connection)
        return XtbImportSummary(
            connection=connection,
            rows_read=sum(item.rows_read for item in parsed_files),
            transactions_added=len(new_transactions),
            duplicates_skipped=len(transactions) - len(new_transactions),
            positions_imported=positions_imported,
            warnings=warnings,
        )


def _resolve_fields(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers if header}
    return {
        field: original
        for field, aliases in HEADER_ALIASES.items()
        if (original := next((normalized[item] for item in aliases if item in normalized), None))
    }


def _normalize_header(value: object) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = "".join(
        character for character in normalized if not unicodedata.combining(character)
    )
    return re.sub(r"[^a-z0-9]", "", ascii_value.lower())


def _value(row: dict[str, object], fields: dict[str, str], field: str) -> str:
    header = fields.get(field)
    value = row.get(header) if header else None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc).isoformat()
    return str(value or "").strip()


def _asset_type(value: str) -> AssetType:
    normalized = _normalize_header(value)
    if "etf" in normalized or "exchangetradedfund" in normalized:
        return AssetType.ETF
    if any(token in normalized for token in ("stock", "share", "equity")):
        return AssetType.STOCK
    if re.search(r"(^|[^a-z])stc([^a-z]|$)", value.lower()):
        return AssetType.STOCK
    return AssetType.OTHER


def _canonical_symbol(symbol: str) -> str:
    return _EXCHANGE_SUFFIX.sub("", symbol.strip().upper())


def _clean_isin(value: str) -> str | None:
    normalized = value.strip().upper()
    return normalized if re.fullmatch(r"[A-Z]{2}[A-Z0-9]{9}[0-9]", normalized) else None


def _row_currency(row: dict[str, object], fields: dict[str, str]) -> str | None:
    explicit = _value(row, fields, "currency").upper()
    if re.fullmatch(r"[A-Z]{3}", explicit):
        return explicit
    for field in ("current_value", "profit", "current_price", "open_price"):
        header = fields.get(field, "")
        raw_value = _value(row, fields, field)
        if currency := _currency_from_text(f"{header} {raw_value}"):
            return currency
    return None


def _report_currency(rows: list[list[object]], filename: str) -> str | None:
    for index, row in enumerate(rows[:30]):
        headers = [_normalize_header(value) for value in row]
        if "currency" not in headers:
            continue
        currency_index = headers.index("currency")
        for candidate_row in rows[index + 1 : index + 8]:
            if currency_index >= len(candidate_row):
                continue
            candidate = str(candidate_row[currency_index] or "").strip().upper()
            if re.fullmatch(r"[A-Z]{3}", candidate):
                return candidate

    filename_currency = re.match(r"^([A-Z]{3})[_-]", Path(filename).name.upper())
    return filename_currency.group(1) if filename_currency else None


def _currency_from_text(value: str) -> str | None:
    if "€" in value:
        return "EUR"
    if "$" in value:
        return "USD"
    if "£" in value:
        return "GBP"
    match = re.search(r"\b([A-Z]{3})\b", value.upper())
    return match.group(1) if match and match.group(1) in _KNOWN_CURRENCIES else None


def _decimal(value: str) -> Decimal | None:
    cleaned = value.strip()
    if not cleaned or cleaned.upper() in {"-", "—", "N/A", "NONE"}:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = re.sub(r"[^0-9,.-]", "", cleaned.strip("()"))
    if not cleaned or cleaned in {"-", ".", ","}:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        parts = cleaned.split(",")
        cleaned = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) <= 4 else "".join(parts)
    try:
        result = Decimal(cleaned)
    except InvalidOperation as exc:
        raise XtbImportError(f"'{value}' is not a valid number") from exc
    return -result if negative else result


def _optional_datetime(value: str) -> datetime | None:
    if not value.strip():
        return None
    normalized = value.strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        parsed = None
        for pattern in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d.%m.%Y",
        ):
            try:
                parsed = datetime.strptime(normalized, pattern)
                break
            except ValueError:
                continue
        if parsed is None:
            raise XtbImportError(f"'{value}' is not a recognized date") from None
    return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def _latest_date(rows: list[list[object]]) -> datetime | None:
    candidates: list[datetime] = []
    for row in rows[:30]:
        row_text = " ".join(str(value or "") for value in row)
        metadata_row = any(
            token in row_text.lower() for token in ("generated", "report date", "as of")
        )
        for value in row:
            if isinstance(value, datetime):
                candidates.append(value if value.tzinfo else value.replace(tzinfo=timezone.utc))
            elif isinstance(value, date):
                candidates.append(datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc))
            elif isinstance(value, str) and metadata_row:
                for match in re.findall(
                    r"\d{4}-\d{2}-\d{2}(?:[ T]\d{2}:\d{2}(?::\d{2})?)?|"
                    r"\d{2}[./]\d{2}[./]\d{4}(?: \d{2}:\d{2}(?::\d{2})?)?",
                    value,
                ):
                    if parsed := _optional_datetime(match):
                        candidates.append(parsed)
    return max(candidates) if candidates else None


def _transaction_type(value: str) -> TransactionType | None:
    normalized = value.lower()
    if "stock purchase" in normalized or "stock sale" in normalized:
        # These cash-operation rows have no share quantity or execution price.
        return TransactionType.OTHER
    if "dividend" in normalized:
        return TransactionType.DIVIDEND
    if "deposit" in normalized or "cash in" in normalized:
        return TransactionType.DEPOSIT
    if "withdraw" in normalized or "cash out" in normalized:
        return TransactionType.WITHDRAWAL
    if any(token in normalized for token in ("commission", "fee", "charge", "tax")):
        return TransactionType.FEE
    if "interest" in normalized:
        return TransactionType.DIVIDEND
    if "buy" in normalized:
        return TransactionType.BUY
    if "sell" in normalized:
        return TransactionType.SELL
    return None


def _transaction(
    *,
    row_id: str,
    action: str,
    ticker: str,
    transaction_type: TransactionType,
    quantity: Decimal | None,
    price: Decimal | None,
    value: Decimal,
    currency: str,
    executed_at: datetime,
) -> ParsedXtbTransaction:
    identity = row_id or "|".join(
        (
            ticker,
            action,
            executed_at.isoformat(),
            str(quantity or ""),
            str(price or ""),
            str(value),
            currency,
        )
    )
    digest = hashlib.sha256(identity.encode()).hexdigest()[:32]
    return ParsedXtbTransaction(
        external_id=f"xtb:{action}:{digest}",
        ticker=ticker.upper(),
        transaction_type=transaction_type,
        quantity=quantity,
        price=price,
        value=abs(value),
        currency=currency,
        executed_at=executed_at,
    )


def _deduplicate_transactions(
    items: list[ParsedXtbTransaction],
) -> list[ParsedXtbTransaction]:
    return list({item.external_id: item for item in items}.values())


def _cash_position(rows: list[list[object]]) -> ParsedXtbPosition | None:
    for row in rows[:50]:
        if not row:
            continue
        label = _normalize_header(row[0])
        if label not in {"availablecash", "cash", "cashbalance"}:
            continue
        raw_value = next(
            (str(value) for value in row[1:] if _decimal(str(value or "")) is not None),
            "",
        )
        value = _decimal(raw_value)
        if value is None:
            continue
        currency = _currency_from_text(" ".join(str(value or "") for value in row)) or "EUR"
        return ParsedXtbPosition(
            instrument_id=f"XTB:CASH:{currency}",
            ticker=currency,
            canonical_symbol=currency,
            name=f"Cash ({currency})",
            isin=None,
            asset_type=AssetType.CASH,
            quantity=value,
            average_price=Decimal(1),
            current_value=value,
            currency=currency,
            reported_pnl=None,
        )
    return None


def _aggregate_positions(items: list[ParsedXtbPosition]) -> list[ParsedXtbPosition]:
    grouped: dict[tuple[str, str], list[ParsedXtbPosition]] = {}
    for item in items:
        grouped.setdefault((item.isin or item.canonical_symbol, item.currency), []).append(item)

    result: list[ParsedXtbPosition] = []
    for group in grouped.values():
        first = group[0]
        asset_type = next(
            (item.asset_type for item in group if item.asset_type is not AssetType.OTHER),
            AssetType.OTHER,
        )
        total_quantity = sum((item.quantity for item in group), Decimal(0))
        cost = sum(
            (
                item.quantity * item.average_price
                for item in group
                if item.average_price is not None
            ),
            Decimal(0),
        )
        priced_quantity = sum(
            (item.quantity for item in group if item.average_price is not None), Decimal(0)
        )
        reported = [item.reported_pnl for item in group if item.reported_pnl is not None]
        result.append(
            ParsedXtbPosition(
                instrument_id=first.instrument_id,
                ticker=first.ticker,
                canonical_symbol=first.canonical_symbol,
                name=first.name,
                isin=first.isin,
                asset_type=asset_type,
                quantity=total_quantity,
                average_price=cost / priced_quantity if priced_quantity else None,
                current_value=sum((item.current_value for item in group), Decimal(0)),
                currency=first.currency,
                reported_pnl=sum(reported, Decimal(0)) if reported else None,
            )
        )
    return sorted(result, key=lambda item: item.current_value, reverse=True)


def _unique(items: list[str]) -> list[str]:
    return list(dict.fromkeys(items))
