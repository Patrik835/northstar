import io
import re
import uuid
import zipfile
from datetime import date
from decimal import Decimal
from typing import Any

import pytest
from openpyxl import Workbook
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session, sessionmaker

import app.models  # noqa: F401
from app.core.database import Base
from app.core.encryption import CredentialCipher
from app.integrations.market_data import FxRateError
from app.models.enums import AssetType, Broker, TransactionType
from app.models.portfolio import Position, Transaction
from app.models.user import User
from app.services.xtb_import import XtbImportService, XtbStatementParser, XtbUpload


class CurrentOnlyRates:
    async def convert_to_eur(
        self, value: Decimal, currency: str, as_of: date | None = None
    ) -> Decimal:
        if as_of is not None:
            raise FxRateError("No historical rate")
        return value


class AsyncSessionAdapter:
    def __init__(self, session: Session) -> None:
        self.session = session

    def add(self, item: Any) -> None:
        self.session.add(item)

    def add_all(self, items: list[Any]) -> None:
        self.session.add_all(items)

    async def flush(self) -> None:
        self.session.flush()

    async def commit(self) -> None:
        self.session.commit()

    async def rollback(self) -> None:
        self.session.rollback()

    async def refresh(self, item: Any) -> None:
        self.session.refresh(item)

    async def execute(self, statement: Any):
        return self.session.execute(statement)

    async def scalar(self, statement: Any):
        return self.session.scalar(statement)

    async def scalars(self, statement: Any):
        return self.session.scalars(statement)

    async def get(self, model: type, item_id: Any):
        return self.session.get(model, item_id)


OPEN_POSITIONS = (
    b"Report generated,2026-03-05 17:30:00\n"
    b"Cash balance,125 EUR\n"
    b"Open positions\n"
    b"Position,Symbol,Instrument type,Name,ISIN,Volume,Open time,Open price,"
    b"Market price,Market value,Currency,Profit\n"
    b"101,AAPL.US,Stock,Apple Inc.,US0378331005,2,2025-01-02 10:00:00,"
    b"180,210,420,EUR,60\n"
    b"102,VUAA.UK,ETF,Vanguard S&P 500 ETF,IE00BFMXXD54,3,"
    b"2025-02-03 11:00:00,90,100,300,EUR,30\n"
)

HISTORY = (
    b"Closed positions\n"
    b"Position,Symbol,Instrument type,Volume,Open time,Open price,Close time,"
    b"Close price,Currency,Profit,Commission\n"
    b"101,AAPL.US,Stock,2,2025-01-02 10:00:00,180,2026-03-06 10:00:00,"
    b"215,EUR,70,-1\n\n"
    b"Cash operations\n"
    b"ID,Type,Date,Value,Currency\n"
    b"cash-1,Dividend,2026-02-01 09:00:00,5,EUR\n"
)


def test_xtb_parser_reads_current_stock_etf_values_and_pnl() -> None:
    parsed = XtbStatementParser().parse(XtbUpload("open-positions.csv", OPEN_POSITIONS))

    assert parsed.has_position_snapshot is True
    assert parsed.valued_at.isoformat().startswith("2026-03-05T17:30:00")
    assert {position.asset_type for position in parsed.positions} == {
        AssetType.STOCK,
        AssetType.ETF,
        AssetType.CASH,
    }
    apple = next(position for position in parsed.positions if position.ticker == "AAPL.US")
    assert apple.canonical_symbol == "AAPL"
    assert apple.current_value == Decimal("420")
    assert apple.reported_pnl == Decimal("60")
    assert parsed.transactions[0].transaction_type is TransactionType.BUY


def test_xtb_parser_supports_excel_and_history_activity() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Closed positions"
    for row in (
        (
            "Position",
            "Symbol",
            "Instrument type",
            "Volume",
            "Open time",
            "Open price",
            "Close time",
            "Close price",
            "Currency",
            "Profit",
        ),
        (
            "301",
            "MSFT.US",
            "Stock",
            1,
            "2025-01-01 10:00:00",
            300,
            "2026-01-01 10:00:00",
            350,
            "EUR",
            50,
        ),
    ):
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)

    parsed = XtbStatementParser().parse(XtbUpload("history.xlsx", buffer.getvalue()))

    assert parsed.has_position_snapshot is False
    assert [item.transaction_type for item in parsed.transactions] == [
        TransactionType.BUY,
        TransactionType.SELL,
    ]


def test_xtb_parser_reads_native_three_sheet_workbook_with_broken_dimensions() -> None:
    workbook = Workbook()
    closed = workbook.active
    closed.title = "Closed Positions"
    closed.append(["Account number", "redacted"])
    closed.append(["Closed Positions"])
    closed.append(
        [
            "Instrument",
            "Ticker",
            "Category",
            "Type",
            "Volume",
            "Open Price",
            "Open Time (UTC)",
            "Close Price",
            "Close Time (UTC)",
            "Profit/Loss",
            "Position ID",
        ]
    )

    cash = workbook.create_sheet("Cash Operations")
    cash.append(["Account number", "redacted"])
    cash.append(["Cash Operations"])
    cash.append(
        [
            "Type",
            "Instrument",
            "Ticker",
            "Category",
            "Time",
            "Amount",
            "ID",
            "Comment",
            "Product",
            "Position ID",
        ]
    )
    cash.append(
        [
            "Dividend",
            "Alphabet",
            "GOOGL.US",
            "STOCK",
            "2026-06-15 05:51:54.820000",
            "0.13",
            "dividend-1",
            "GOOGL.US USD dividend",
            "My Trades",
            "position-2",
        ]
    )
    cash.append(
        [
            "Withholding tax",
            "Alphabet",
            "GOOGL.US",
            "STOCK",
            "2026-06-15 05:51:54.820000",
            "-0.04",
            "tax-1",
            "GOOGL.US USD WHT",
            "My Trades",
            "position-2",
        ]
    )
    cash.append(
        [
            "Stock purchase",
            "Alphabet",
            "GOOGL.US",
            "STOCK",
            "2025-03-10 14:46:25.753000",
            "-165",
            "purchase-1",
            "Stock purchase",
            "My Trades",
            "position-2",
        ]
    )
    cash.append(
        [
            "Free funds interest",
            "",
            "",
            "",
            "2026-07-01 00:00:00",
            "0.10",
            "interest-1",
            "Free funds interest USD",
            "My Trades",
            "",
        ]
    )

    opened = workbook.create_sheet("Open Positions")
    opened.append(["Account number", "redacted"])
    opened.append(["Open Positions"])
    opened.append(["Data as of report generated", "2026-08-08 19:20:20.647000"])
    opened.append(["Product", "Metric", "Amount", "Currency"])
    opened.append(["My Trades", "Value", "714", "USD"])
    opened.append([])
    opened.append(
        [
            "Product",
            "Instrument/Position",
            "Ticker",
            "Category",
            "Type",
            "Volume",
            "Value",
            "Current price",
            "Open price",
            "Open time (UTC)",
            "Net Profit %",
            "Net Profit",
        ]
    )
    opened.append(
        ["My Trades", "AMD", "AMD.US", "STOCK", "", "1", "500", "", "100", "", "400", "400"]
    )
    opened.append(
        [
            "My Trades",
            "position-1",
            "AMD.US",
            "",
            "BUY",
            "1",
            "500",
            "500",
            "100",
            "2025-03-04 15:18:04.012000",
            "400",
            "400",
        ]
    )
    opened.append(
        ["My Trades", "Alphabet", "GOOGL.US", "STOCK", "", "1", "214", "", "165", "", "30", "49"]
    )
    opened.append(
        [
            "My Trades",
            "position-2",
            "GOOGL.US",
            "",
            "BUY",
            "1",
            "214",
            "214",
            "165",
            "2025-03-10 14:46:25.753000",
            "30",
            "49",
        ]
    )

    source = io.BytesIO()
    workbook.save(source)
    files: dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(source.getvalue())) as archive:
        for info in archive.infolist():
            content = archive.read(info.filename)
            if info.filename.startswith("xl/worksheets/sheet"):
                content = re.sub(rb'<dimension ref="[^"]+"', b'<dimension ref="A1:A1"', content)
            files[info.filename] = content
    native = io.BytesIO()
    with zipfile.ZipFile(native, "w") as archive:
        for filename, content in files.items():
            archive.writestr(filename, content)

    parsed = XtbStatementParser().parse(
        XtbUpload("USD_redacted_2006-01-01_2026-08-08.xlsx", native.getvalue())
    )

    assert len(parsed.positions) == 2
    assert {item.name for item in parsed.positions} == {"AMD", "Alphabet"}
    assert {item.currency for item in parsed.positions} == {"USD"}
    assert {item.reported_pnl for item in parsed.positions} == {
        Decimal("400"),
        Decimal("49"),
    }
    assert [item.transaction_type for item in parsed.transactions].count(TransactionType.BUY) == 2
    assert [item.transaction_type for item in parsed.transactions].count(
        TransactionType.DIVIDEND
    ) == 2
    assert [item.transaction_type for item in parsed.transactions].count(TransactionType.FEE) == 1
    assert [item.transaction_type for item in parsed.transactions].count(TransactionType.OTHER) == 1


@pytest.mark.asyncio
async def test_history_only_reimport_preserves_positions_and_deduplicates_activity() -> None:
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(engine, expire_on_commit=False)

    user_id = uuid.uuid4()
    try:
        with session_factory() as sync_db:
            db = AsyncSessionAdapter(sync_db)
            db.add(
                User(
                    id=user_id,
                    username="xtb-owner",
                    password_hash="unused",
                    is_admin=False,
                    is_active=True,
                )
            )
            await db.commit()
            service = XtbImportService(
                db,  # type: ignore[arg-type]
                CredentialCipher("MDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDA="),
                fx_rates=CurrentOnlyRates(),
            )

            first = await service.import_files(
                user_id, [XtbUpload("open-positions.csv", OPEN_POSITIONS)]
            )
            second = await service.import_files(user_id, [XtbUpload("history.csv", HISTORY)])
            third = await service.import_files(user_id, [XtbUpload("history.csv", HISTORY)])

            assert first.positions_imported == 3
            assert any("no matching stored ECB rate" in item for item in first.warnings)
            assert second.positions_imported == 0
            assert third.transactions_added == 0
            assert third.duplicates_skipped == 4
            assert await db.scalar(select(func.count(Position.id))) == 3
            assert await db.scalar(select(func.count(Transaction.id))) == 5
            connection = first.connection
            assert connection.broker is Broker.XTB
            apple = await db.scalar(select(Position).where(Position.ticker == "AAPL.US"))
            assert apple is not None
            assert apple.reported_pnl_eur == Decimal("60")
    finally:
        engine.dispose()
